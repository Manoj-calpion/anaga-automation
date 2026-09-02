"""CLI: batch loop, circuit breaker, resume, roster vs per-license."""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from playwright.async_api import TimeoutError as PlaywrightTimeout

from browser import BrowserSession, dismiss_css_error
from detail import numbers_match, save_page_pdf, scrape_detail, wait_for_detail
from logging_ import RunWriters
from models import (
    CloudflareChallenge,
    CssErrorModal,
    INACTIVE_STATUSES,
    InputRow,
    RecaptchaCircuitOpen,
    RunLogRow,
    SearchHit,
    StatusCode,
)
from naming import PathTooLongError, resolve_pdf_path
from search import (
    AuraTap,
    RecaptchaFailed,
    SearchValidationError,
    click_search,
    click_select_for_license,
    fill_search_form,
    parse_results_from_dom,
    wait_for_results,
)

BACKOFFS = (5, 15, 45)


def step(msg: str) -> None:
    print(f"  {msg}", flush=True)


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def load_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as fh:
        data = yaml.safe_load(fh)
    if not isinstance(data, dict):
        raise SystemExit(f"Config is empty or invalid: {path}")
    return data


def load_input_rows(config: dict[str, Any]) -> list[InputRow]:
    path = Path(config["input_file"])
    col = config.get("input_column", "license_number")
    type_col = config.get("license_type_column") or "license_type"
    default_type = config.get("license_type_default", "Behavior Analyst")
    if not path.is_absolute():
        path = Path.cwd() / path
    if not path.exists():
        raise SystemExit(f"Input file not found: {path}")

    raw_rows: list[dict[str, Any]] = []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as fh:
            raw_rows = list(csv.DictReader(fh))
    elif suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            rec = {headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(excel_row) if i < len(headers)}
            if any(rec.values()):
                raw_rows.append(rec)
    else:
        raise SystemExit(f"Unsupported input type {suffix}; use .csv or .xlsx")

    def pick(raw: dict[str, Any], name: str) -> str:
        for k, v in raw.items():
            if k and str(k).strip().lower() == name.strip().lower():
                return "" if v is None else str(v).strip()
        return ""

    items: list[InputRow] = []
    for raw in raw_rows:
        items.append(
            InputRow(
                license_number=pick(raw, col),
                license_type=pick(raw, type_col) or default_type,
                source={str(k): ("" if v is None else str(v)) for k, v in raw.items()},
            )
        )
    return items


def log_row(
    license_number: str,
    code: StatusCode,
    *,
    error: str = "",
    pdf_path: str = "",
    record=None,
    extra: dict[str, str] | None = None,
) -> RunLogRow:
    fields = record.as_log_fields() if record is not None else {}
    return RunLogRow(
        license_number=license_number,
        requested_at=utc_now(),
        status_code=code.value,
        provider_name=fields.get("provider_name", ""),
        license_type=fields.get("license_type", ""),
        license_status=fields.get("license_status", ""),
        issued_date=fields.get("issued_date", ""),
        expiration_date=fields.get("expiration_date", ""),
        pdf_path=pdf_path,
        error_detail=error,
        extra=extra or {},
    )


async def backoff(attempt: int) -> None:
    seconds = BACKOFFS[min(attempt, len(BACKOFFS) - 1)]
    print(f"Backoff {seconds}s")
    await asyncio.sleep(seconds)


def captures_file(output_root: Path) -> Path:
    return output_root / ".captures.json"


def load_captures(output_root: Path) -> dict[str, str]:
    path = captures_file(output_root)
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return {str(k).upper(): str(v) for k, v in data.items()}


def save_captures(output_root: Path, captures: dict[str, str]) -> None:
    path = captures_file(output_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(captures, indent=2), encoding="utf-8")


def already_logged_licenses(writers: RunWriters) -> set[str]:
    path = writers.run_log_path
    if not path.exists():
        return set()
    seen: set[str] = set()
    with path.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            lic = (row.get("license_number") or "").strip()
            if lic:
                seen.add(lic.upper())
    return seen


async def perform_search(
    session: BrowserSession,
    tap: AuraTap,
    *,
    license_number: str,
    license_type: str,
) -> tuple[list[SearchHit], int | None]:
    config = session.config
    page = session.page
    assert page
    tap.bodies.clear()
    step("Opening the Georgia license search page...")
    await session.goto_search()
    step(f"Filling the form for {license_number or '(roster search)'}...")
    await fill_search_form(
        page,
        profession=config.get("profession_type", "Behavior Analyst"),
        license_type=license_type,
        license_number=license_number,
        ui_pause=session.ui_pause,
    )
    from search import form_matches

    if not await form_matches(
        page,
        profession=config.get("profession_type", "Behavior Analyst"),
        license_type=license_type,
        license_number=license_number,
    ):
        await fill_search_form(
            page,
            profession=config.get("profession_type", "Behavior Analyst"),
            license_type=license_type,
            license_number=license_number,
            ui_pause=session.ui_pause,
        )
    step("Clicking Search...")
    await click_search(page, session.ui_pause, mode=config.get("search_click_mode", "auto"))
    step("Waiting for results...")
    try:
        await wait_for_results(page, timeout_ms=int(config.get("interaction_timeout_ms", 30000)))
    except PlaywrightTimeout:
        hits, total, recaptcha = tap.latest_search()
        if recaptcha:
            raise RecaptchaFailed("V3 Recaptcha failed in apex")
        if hits:
            return hits, total
        raise
    hits, total, recaptcha = tap.latest_search()
    if recaptcha:
        raise RecaptchaFailed("V3 Recaptcha failed in apex")
    if not hits:
        hits = await parse_results_from_dom(page)
        total = total if total is not None else len(hits)
    return hits, total


async def save_record_pdf(
    session: BrowserSession,
    writers: RunWriters,
    item: InputRow,
    rec,
    occupied: set[str],
) -> RunLogRow:
    config = session.config
    output_root = Path(config["output_root"])
    quarantine = Path(config.get("quarantine_folder", str(output_root / "quarantine")))
    extra: dict[str, str] = {}
    if rec.status.strip().lower() in INACTIVE_STATUSES:
        extra["status_flag"] = rec.status
        finding = f"finding: license status {rec.status}"
    else:
        finding = ""

    try:
        dest, expiry_missing, _note = resolve_pdf_path(
            output_root,
            first_name=rec.first_name,
            middle_name=rec.middle,
            last_name=rec.last_name,
            state_code=config.get("state_code", "GA"),
            license_type=rec.license_type,
            expires=rec.expires,
            date_format=config.get("date_format_in_filename", "MM-DD-YYYY"),
            separator=config.get("filename_separator", " - "),
            license_number=item.license_number,
            output_layout=config.get("output_layout", "flat"),
            on_existing_file=config.get("on_existing_file", "skip"),
            occupied=occupied,
        )
    except PathTooLongError as exc:
        dest = quarantine / f"{item.license_number}.pdf"
        expiry_missing = True
        finding = (finding + "; " if finding else "") + f"path too long, quarantine: {exc}"

    on_existing = config.get("on_existing_file", "skip")
    if dest.exists() and on_existing == "skip":
        occupied.add(str(dest.resolve()))
        return log_row(
            item.license_number,
            StatusCode.SKIPPED_EXISTS,
            pdf_path=str(dest),
            record=rec,
            extra=extra,
        )

    page = session.page
    assert page
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        await save_page_pdf(page, dest)
    except Exception as exc:
        quarantine.mkdir(parents=True, exist_ok=True)
        qdest = quarantine / dest.name
        try:
            await save_page_pdf(page, qdest)
            dest = qdest
            finding = (finding + "; " if finding else "") + f"wrote quarantine ({exc})"
        except Exception as exc2:
            return log_row(
                item.license_number,
                StatusCode.ERROR,
                error=f"PDF unwritable: {exc}; quarantine failed: {exc2}",
                record=rec,
                extra=extra,
            )

    occupied.add(str(dest.resolve()))
    code = StatusCode.NO_EXPIRY if expiry_missing else StatusCode.OK
    err = finding
    if expiry_missing:
        err = (err + "; " if err else "") + "EXPIRES blank or '-'"
    return log_row(item.license_number, code, error=err, pdf_path=str(dest), record=rec, extra=extra)


async def lookup_per_license(session: BrowserSession, tap: AuraTap, item: InputRow, occupied: set[str], writers: RunWriters) -> RunLogRow:
    hits, total = await perform_search(
        session, tap, license_number=item.license_number, license_type=item.license_type
    )
    n = total if total is not None else len(hits)
    if n == 0 or not hits:
        return log_row(item.license_number, StatusCode.NOT_FOUND, error="0 search results")
    if n > 1 or len(hits) > 1:
        names = [f"{h.license_number} ({h.full_name})" for h in hits]
        return log_row(
            item.license_number,
            StatusCode.AMBIGUOUS,
            error="candidates: " + "; ".join(names),
        )
    hit = hits[0]
    if hit.license_number and not numbers_match(item.license_number, hit.license_number):
        return log_row(
            item.license_number,
            StatusCode.MISMATCH,
            error=f"result {hit.license_number!r} != requested {item.license_number!r}",
        )
    page = session.page
    assert page
    await click_select_for_license(page, item.license_number, session.ui_pause)
    await wait_for_detail(page)
    rec = await scrape_detail(page)
    if not numbers_match(item.license_number, rec.license_number):
        return log_row(
            item.license_number,
            StatusCode.MISMATCH,
            error=f"detail {rec.license_number!r} != requested {item.license_number!r}",
            record=rec,
        )
    return await save_record_pdf(session, writers, item, rec, occupied)


async def open_detail_token(session: BrowserSession, token: str, license_number: str) -> bool:
    from detail import detail_url
    from search import decode_encrypted_id_variants

    page = session.page
    assert page
    base = session.config.get("search_url")
    for variant in decode_encrypted_id_variants(token):
        url = detail_url(base, variant)
        await page.goto(url, wait_until="domcontentloaded")
        try:
            await wait_for_detail(page, timeout_ms=20000)
            rec = await scrape_detail(page)
            if numbers_match(license_number, rec.license_number) or rec.license_number:
                return True
        except Exception:
            continue
    return False


async def lookup_via_token(session: BrowserSession, item: InputRow, token: str, occupied: set[str], writers: RunWriters) -> RunLogRow:
    ok = await open_detail_token(session, token, item.license_number)
    if not ok:
        return log_row(item.license_number, StatusCode.DETAIL_LOAD_FAILED, error="roster token did not render detail")
    rec = await scrape_detail(session.page)
    if not numbers_match(item.license_number, rec.license_number):
        return log_row(
            item.license_number,
            StatusCode.MISMATCH,
            error=f"detail {rec.license_number!r} != requested",
            record=rec,
        )
    return await save_record_pdf(session, writers, item, rec, occupied)


async def collect_roster(session: BrowserSession, tap: AuraTap) -> dict[str, str]:
    hits, total = await perform_search(
        session,
        tap,
        license_number="",
        license_type=session.config.get("license_type_default", "Behavior Analyst"),
    )
    index: dict[str, str] = {}
    for h in hits:
        if h.license_number and h.encrypted_license_id:
            index[h.license_number.strip().upper()] = h.encrypted_license_id
    print(f"Roster intercepted {len(hits)} rows, totalRows={total}")
    page = session.page
    assert page
    # Pagination is not CAPTCHA-gated in the spec; walk Next if present.
    for _ in range(50):
        nxt = page.get_by_role("button", name="Next")
        if await nxt.count() == 0:
            break
        try:
            if await nxt.first.is_disabled():
                break
        except Exception:
            break
        tap.bodies.clear()
        await session.ui_pause()
        await nxt.first.click()
        await wait_for_results(page)
        more, _, recaptcha = tap.latest_search()
        if recaptcha:
            raise RecaptchaFailed("V3 Recaptcha failed in apex")
        if not more:
            more = await parse_results_from_dom(page)
        added = 0
        for h in more:
            key = h.license_number.strip().upper()
            if key and h.encrypted_license_id and key not in index:
                index[key] = h.encrypted_license_id
                added += 1
        if added == 0:
            break
    return index


async def process_item_with_retries(
    session: BrowserSession,
    tap: AuraTap,
    item: InputRow,
    occupied: set[str],
    writers: RunWriters,
    *,
    roster: dict[str, str] | None,
    consecutive_recaptcha: list[int],
) -> RunLogRow:
    config = session.config
    recaptcha_tries = 0
    hang_retried = False
    css_retried = False
    detail_retried = False
    max_consecutive = int(config.get("recaptcha_max_consecutive", 3))

    while True:
        try:
            await dismiss_css_error(session.page)
            if roster is not None:
                token = roster.get(item.license_number.strip().upper())
                if not token:
                    return log_row(item.license_number, StatusCode.NOT_FOUND, error="not present in roster search")
                row = await lookup_via_token(session, item, token, occupied, writers)
            else:
                row = await lookup_per_license(session, tap, item, occupied, writers)
            consecutive_recaptcha[0] = 0
            return row
        except RecaptchaFailed as exc:
            recaptcha_tries += 1
            consecutive_recaptcha[0] += 1
            if consecutive_recaptcha[0] >= max_consecutive:
                raise RecaptchaCircuitOpen(str(exc)) from exc
            if recaptcha_tries >= 3:
                raise RecaptchaCircuitOpen("recaptcha retries exhausted") from exc
            await backoff(recaptcha_tries - 1)
        except CloudflareChallenge:
            raise
        except CssErrorModal:
            if not css_retried:
                css_retried = True
                print("  Lightning CSS error — returning to search without a hard refresh.")
                await session.goto_search()
                continue
            return log_row(item.license_number, StatusCode.ERROR, error="CSS Error modal persisted")
        except SearchValidationError as exc:
            if not hang_retried:
                hang_retried = True
                print(f"  Dropdown did not take the value ({exc}). Retrying the form...")
                await asyncio.sleep(2)
                continue
            print(f"  ERROR: {exc}")
            return log_row(item.license_number, StatusCode.ERROR, error=str(exc))
        except (PlaywrightTimeout, RuntimeError) as exc:
            msg = str(exc).lower()
            if "detail" in msg and not detail_retried:
                detail_retried = True
                print("  Detail did not load; retrying without a page refresh.")
                continue
            if not hang_retried:
                hang_retried = True
                print("  Retrying on the same page (not refreshing — that would redo Cloudflare)...")
                continue
            code = StatusCode.DETAIL_LOAD_FAILED if "detail" in msg else StatusCode.ERROR
            return log_row(item.license_number, code, error=str(exc))


async def run_batch(config: dict[str, Any], *, fresh_log: bool) -> int:
    items = load_input_rows(config)
    output_root = Path(config["output_root"])
    if not output_root.is_absolute():
        output_root = Path.cwd() / output_root
    config["output_root"] = str(output_root)
    q = Path(config.get("quarantine_folder", str(output_root / "quarantine")))
    if not q.is_absolute():
        q = Path.cwd() / q
    config["quarantine_folder"] = str(q)
    writers = RunWriters(output_root, config.get("input_column", "license_number"))
    if fresh_log:
        for p in (writers.run_log_path, writers.failures_path):
            if p.exists():
                p.unlink()

    prior_ok = writers.load_completed_ok_licenses()
    captures = load_captures(output_root)
    for lic, pdf in captures.items():
        if Path(pdf).exists():
            prior_ok.setdefault(lic, pdf)
            prior_ok.setdefault(lic.upper(), pdf)
    occupied: set[str] = {str(Path(p).resolve()) for p in prior_ok.values() if Path(p).exists()}
    on_existing = config.get("on_existing_file", "skip")
    lookup_mode = config.get("lookup_mode", "per_license")
    consecutive_recaptcha = [0]
    halt_reason = ""
    print(f"Loaded {len(items)} license numbers from {config['input_file']}")
    print("A Chrome window will open.")
    print("If you see a 'verify you are human' checkbox: click it ONCE, then wait.")
    print("Do not click it again. Do not refresh. Do not click Search.")
    print("Clicking the box many times makes the site show it again.")
    if config.get("search_click_mode") == "human":
        print("HUMAN MODE is on: the script will pause and ask YOU to click Search.")
        print("For hands-off automation, stop this run and start without --human-search-click.")
    print()

    async with BrowserSession(config) as session:
        tap = AuraTap(session.page)
        tap.start()
        roster: dict[str, str] | None = None
        if lookup_mode == "roster":
            try:
                await session.rate.acquire()
                roster = await collect_roster(session, tap)
            except RecaptchaFailed as exc:
                print(f"Roster search failed recaptcha ({exc}); falling back to per_license")
                roster = None
            except RecaptchaCircuitOpen:
                raise

        remaining_halt = False
        for i, item in enumerate(items):
            if remaining_halt:
                writers.append(
                    log_row(item.license_number, StatusCode.HALTED, error=halt_reason),
                    item.original_columns(),
                )
                continue
            if not item.license_number:
                writers.append(log_row("", StatusCode.ERROR, error="blank license_number"), item.original_columns())
                continue

            key = item.license_number.upper()
            latest = writers.latest_by_license().get(key)
            latest_code = (latest or {}).get("status_code", "")
            # Only skip terminal successes. HALTED / ERROR / NOT_FOUND must retry.
            skip_codes = {
                StatusCode.OK.value,
                StatusCode.SKIPPED_EXISTS.value,
                StatusCode.NO_EXPIRY.value,
            }
            if (
                not fresh_log
                and on_existing == "skip"
                and latest_code in skip_codes
            ):
                pdf = (latest or {}).get("pdf_path") or prior_ok.get(item.license_number) or prior_ok.get(key) or ""
                if latest_code == StatusCode.SKIPPED_EXISTS.value or (pdf and Path(pdf).exists()):
                    print(f"[{i + 1}/{len(items)}] {item.license_number} {latest_code} — skip")
                    continue
            if on_existing == "skip" and (
                item.license_number in prior_ok or item.license_number.upper() in prior_ok
            ) and not latest:
                pdf = prior_ok.get(item.license_number) or prior_ok.get(item.license_number.upper())
                writers.append(
                    log_row(
                        item.license_number,
                        StatusCode.SKIPPED_EXISTS,
                        pdf_path=pdf or "",
                    ),
                    item.original_columns(),
                )
                print(f"[{i + 1}/{len(items)}] {item.license_number} SKIPPED_EXISTS")
                continue

            print(f"[{i + 1}/{len(items)}] Looking up {item.license_number} ...")
            await session.rate.acquire()
            try:
                row = await process_item_with_retries(
                    session,
                    tap,
                    item,
                    occupied,
                    writers,
                    roster=roster,
                    consecutive_recaptcha=consecutive_recaptcha,
                )
            except RecaptchaCircuitOpen as exc:
                remaining_halt = True
                halt_reason = f"circuit breaker: {exc}"
                print(halt_reason)
                writers.append(log_row(item.license_number, StatusCode.HALTED, error=halt_reason), item.original_columns())
                continue
            except CloudflareChallenge as exc:
                remaining_halt = True
                halt_reason = f"Cloudflare challenge: {exc}"
                print(halt_reason)
                writers.append(log_row(item.license_number, StatusCode.HALTED, error=halt_reason), item.original_columns())
                continue
            except Exception as exc:
                import traceback

                err = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()}"
                print(err)
                writers.append(
                    log_row(item.license_number, StatusCode.ERROR, error=err),
                    item.original_columns(),
                )
                continue

            writers.append(row, item.original_columns())
            if row.pdf_path and row.status_code in {StatusCode.OK.value, StatusCode.NO_EXPIRY.value}:
                captures[item.license_number.upper()] = row.pdf_path
                save_captures(output_root, captures)
            print(f"[{i + 1}/{len(items)}] {item.license_number} {row.status_code}"
                  + (f" — {row.error_detail}" if row.error_detail else ""))
            if i < len(items) - 1 and not remaining_halt:
                await session.provider_pause()

        tap.stop()

    writers.print_summary(expected_rows=len(items))
    final_count = writers.existing_row_count()
    if final_count != len(items):
        print(
            f"WARNING: run_log has {final_count} rows, input has {len(items)}. "
            "Re-run with --resume (default) after a crash, or --fresh to start a new log."
        )
    if halt_reason:
        print(f"RUN HALTED: {halt_reason}")
        return 2
    return 0


def cli(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Georgia SOS Behavior Analyst license lookup")
    parser.add_argument("--config", default="config.yaml")
    parser.add_argument("--spike", action="store_true", help="Run Phase 0 roster/token spike")
    parser.add_argument("--fresh", action="store_true", help="Start a new run_log.csv / failures.csv")
    parser.add_argument(
        "--human-search-click",
        action="store_true",
        help="Optional. Pause so YOU click Search. Default is fully automatic (no captcha puzzle).",
    )
    parser.add_argument("--lookup-mode", choices=["per_license", "roster"], help="Override config lookup_mode")
    args = parser.parse_args(argv)
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = Path.cwd() / config_path
    config = load_yaml(config_path)
    if args.human_search_click:
        config["search_click_mode"] = "human"
    if args.lookup_mode:
        config["lookup_mode"] = args.lookup_mode
    if args.spike:
        from spike_runner import run_spike

        raise SystemExit(asyncio.run(run_spike(config)))
    raise SystemExit(asyncio.run(run_batch(config, fresh_log=args.fresh)))


def main() -> None:
    cli()


if __name__ == "__main__":
    main()
